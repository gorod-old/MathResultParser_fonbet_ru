# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import math
import os
import re
import sys
import threading
import winsound
from random import uniform, choice
from time import sleep, perf_counter
from datetime import datetime
import subprocess

from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem, QApplication, QMainWindow, QWidget, QGridLayout, QSlider
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from selenium import webdriver
import selenium.webdriver.chrome.service as service
from selenium.webdriver import DesiredCapabilities
from openpyxl import Workbook

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt, QModelIndex, QSize
from bs4 import BeautifulSoup as BS

app_window = None
host = 'https://www.fonbet.ru/live'
ajax_params = {}
user_agents_list = []
proxies_list = []
not_parse = []
encoding = 'utf-8'


def beep():
    if app_window and app_window.sound:
        frequency = 2500  # Set Frequency To 2500 Hertz
        duration = 500  # Set Duration To 1000 ms == 1 second
        winsound.Beep(frequency, duration)


def time_str(num):
    if num < 10:
        return '0' + str(num)
    return str(num)


def get_time(sec):
    hour_ = math.trunc(sec / 3600)
    min_ = math.trunc(sec % 3600 / 60)
    sec_ = math.trunc(sec % 3600 % 60)
    return time_str(hour_) + ':' + time_str(min_) + ':' + time_str(sec_)


def get_user_agents_list():
    ua_list = open('user-agents.txt').read().strip().split('\n')
    for ua in ua_list:
        if len(ua) == 0:
            ua_list.remove(ua)
    return ua_list


def get_proxies_list():
    p_list = open('proxies.txt').read().strip().split('\n')
    for p in p_list:
        if len(p) == 0:
            p_list.remove(p)
    return p_list


def save_html(html_str):
    html = open("page.html", "w", encoding=encoding)
    html.seek(0)
    html.write(html_str)
    html.close()


def write_xlsx_data(start_time, p_data):
    if len(p_data) == 0:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Items"
    ws.row_dimensions[2].height = 60
    # row color
    fill_1 = PatternFill(start_color='f5eeda',
                         end_color='f5eeda',
                         fill_type='solid')
    # red color
    fill_2 = PatternFill(start_color='f5c7bf',
                         end_color='f5c7bf',
                         fill_type='solid')
    # header color
    fill_3 = PatternFill(start_color='e8ffdb',
                         end_color='e8ffdb',
                         fill_type='solid')
    # green
    fill_4 = PatternFill(start_color='adf7c1',
                         end_color='adf7c1',
                         fill_type='solid')
    # orange
    fill_5 = PatternFill(start_color='fce079',
                         end_color='fce079',
                         fill_type='solid')
    n = 3  # start row number without header (with a report from 1)
    row = n
    col = 1
    for data_row in p_data:
        for key in data_row.keys():
            if row == n or key not in p_data[row - (n + 1)].keys():
                row_ = row
                check = False
                for index in range(row - n, len(p_data)):
                    if p_data[index][key][0] != '':
                        check = True
                        for i in range(4):
                            ws.cell(row=row_, column=col + i).value = p_data[index][key][i]
                            if (row_ + 1) % 2 == 0:
                                ws.cell(row=row_, column=col + i).fill = fill_1
                            if i == 1:
                                if '!' in p_data[index][key][i]:
                                    ws.cell(row=row_, column=col + i).fill = fill_5
                            if i == 3:
                                if '+' in p_data[index][key][i]:
                                    ws.cell(row=row_, column=col + i).fill = fill_4
                                else:
                                    ws.cell(row=row_, column=col + i).fill = fill_2
                    row_ += 1
                if check:
                    h_row = n - 1  # header row
                    ws.cell(row=h_row, column=col).value = key
                    ws.column_dimensions[get_column_letter(col)].width = 10
                    for i in range(3):
                        ws.cell(row=h_row, column=col + i).fill = fill_3
                    for index in range(h_row, row):
                        ws.cell(row=index, column=col + 3).fill = fill_2
                    col += 4
        ws.row_dimensions[row].height = 20
        row += 1
    # header info row (start and end time)
    end_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S").replace('/', '-').replace(' ', '_').replace(':', '-')
    ws.cell(row=1, column=1).value = f'Начало парсинга: {start_time}, конец парсинга: {end_time}'
    for i in range(8):
        ws.cell(row=1, column=i + 1).fill = fill_5

    save_xlsx(wb)


def set_xlsx_col_width(ws):
    column_widths = []
    for row in ws.rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell.value) > column_widths[i]:
                    column_widths[i] = len(cell.value)
            else:
                if cell.value:
                    column_widths += [len(cell.value)]
                else:
                    column_widths += [5]

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width


def save_xlsx(wb):
    d_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S").replace('/', '-').replace(' ', '_').replace(':', '-')
    f_name = f'/result data/result-{d_time}.xlsx'
    root = os.getcwd() + '/result data'
    if not os.path.isdir(root):
        os.mkdir(root)
    path = os.getcwd() + f_name
    print('file path:' + path)
    if os.path.exists(path):
        os.remove(path)
    wb.save(path)
    # start xlsx file:
    # os.startfile(path)


def set_variables():
    global user_agents_list, proxies_list, app_window
    if os.path.exists(os.getcwd() + '/user-agents.txt'):
        print('user-agents.txt - is found')
        user_agents_list = get_user_agents_list()
    if os.path.exists(os.getcwd() + '/proxies.txt'):
        print('proxies.txt - is found')
        proxies_list = get_proxies_list()
    if app_window and len(proxies_list) == 0:
        app_window.proxy_off()


def parse_data_selenium(_url):
    # 10 попыток запросов на сервер с временной отсрочкой сменой ip и user-agent
    _driver = None
    for i in range(10):
        try:
            _driver = get_w_driver()
            _driver.get(_url)
        except Exception as e:
            print(str(e))
        sleep(1)
        if _driver and _driver.page_source != '<html><head></head><body></body></html>' \
                and 'Ваш браузер устарел' not in _driver.page_source and 'old-browser' not in _driver.current_url:
            break
    return _driver


def get_w_driver():
    timeout = uniform(0, .2)
    sleep(timeout)
    # prepare the option for the chrome driver
    options = webdriver.ChromeOptions()
    # options.headless = True
    # options.add_argument('user-agent=' + choice(user_agents_list))
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('disable-blink-features=AutomationControlled')
    webdriver.DesiredCapabilities.CHROME['acceptSslCerts'] = True
    if app_window and app_window.proxy and len(proxies_list) > 0:
        prx = choice(proxies_list)
        print('proxy: ', prx)
        options.add_argument('--proxy-server=' + prx)

    # start chrome browser
    path = os.getcwd() + '/chromedriver_win32/chromedriver.exe'
    driver = webdriver.Chrome(executable_path=path, options=options)
    return driver


def get_w_driver_1():
    timeout = uniform(0, .2)
    sleep(timeout)
    # start chrome browser
    path = os.getcwd() + '/chromedriver_win32/chromedriver.exe'
    serv = service.Service(path)
    serv.start()
    # prepare the option for the chrome driver
    options = webdriver.ChromeOptions()
    # options.headless = True
    capabilities = DesiredCapabilities.CHROME
    driver = webdriver.Remote(serv.service_url, desired_capabilities=capabilities, options=options)
    return driver


def get_page_data():
    global app_window
    exception = False
    driver = app_window.driver
    tables = None
    while app_window and app_window.run:
        i = 0
        while not tables or not len(tables) > 0:
            if not app_window.run:
                break
            # if not app_window.run or i >= 60:
            #     break
            sleep(app_window.interval)
            try:
                html = BS(driver.page_source, 'html.parser')
                tables = html.findAll(title=re.compile("FIFA"), class_='table__title-text')
                if len(tables) == 0:
                    # beep()
                    print('tables: ', tables)
            except Exception as e:
                print(str(e))
                exception = True
                break
            finally:
                pass
            i += 1

        row_data = {}
        if tables:
            for tbl in tables:
                block = tbl.find_parent('tbody')
                match_list = block.select('tr.table__row')
                i = 0
                for match in match_list:
                    if i > 0:
                        name = None
                        m_time = None
                        score = None
                        total = None
                        try:
                            name = match.select_one('div.table__match-title-text').text
                            name = name.replace(' — ', '\n — ').replace('(', '\n(')
                            m_time = match.select_one('div.table__time span.table__time-text').text
                            score = match.select_one('div.table__score').text
                            total = match.select_one('td:nth-child(13)').text
                        except Exception as e:
                            # print(str(e))
                            pass
                        if name and m_time and score:
                            match_data = {
                                'name': name,
                                'score': score,
                                'time': m_time,
                                'total': total
                            }
                            row_data.update({name: match_data})
                    i += 1

        if len(row_data) == 0:
            t = threading.Thread(target=app_window.set_num_matches, args=(0,))
            t.start()
            t.join()

        if app_window.run:
            if len(row_data) > 0:
                app_window.add_row_data(row_data)
            elif exception:
                t = threading.Thread(target=app_window.reload_driver, args=('driver reload: error occurred',))
                t.start()
                t.join()
                print('driver is ready')
                driver = app_window.driver
                exception = False
        tables = None


class AsyncProcess:
    def __init__(self, name, function, stream_num, callback, args=()):
        super(AsyncProcess, self).__init__()

        self.stream_list = []
        self.start_time = perf_counter()
        print('process: "', name, '", start time: ', self.start_time)
        self.name = name
        self.stream_num = stream_num
        self.callback = callback
        for num in range(stream_num):
            args_ = args
            if stream_num > 1:
                args_ = args + (num + 1,)  # добавляем номер процесса
            self.stream_list.append(threading.Thread(target=function, args=args_))
            self.stream_list[num].start()
            sleep(1)
        t = threading.Thread(target=self.waiting_for_process_end, args=())
        t.start()

    def waiting_for_process_end(self):
        for num in range(self.stream_num):
            self.stream_list[num].join()
        ov_time = perf_counter() - self.start_time
        print('process: "', self.name, '", end time: ', perf_counter())
        print('process: "', self.name, '" - completed, total time: ', get_time(ov_time), ' sec')
        if app_window and self.callback:
            class_method = getattr(MainWindow, self.callback)
            class_method(app_window)


class MainWindow(QMainWindow):
    # Переопределяем конструктор класса
    def __init__(self):
        # Обязательно нужно вызвать метод супер класса
        QMainWindow.__init__(self)

        # colors
        self.grn_color = QtGui.QColor(173, 247, 193)
        self.red_color = QtGui.QColor(245, 199, 191)
        self.r_color = QtGui.QColor(245, 238, 218)
        self.org_color = QtGui.QColor(252, 224, 121)

        self.start_time = None
        self.sound = True
        self.proxy = False
        self.run = False
        self.end = True
        self.close_app = False
        self.auto = False
        self.driver = None
        self.interval = 1
        self.total_list = {}  # последние значения total по матчам
        self.pars_rows = []  # массив данных спаршеных но не показаных на экран
        self.xlsx_data = []  # data to write xlsx
        self.end_match_check = {}  # number off last empty rows for matches
        self.row_color = 0  # 0 or 1 value permissible

        self.setMinimumSize(QSize(1800, 800))  # Устанавливаем размеры
        self.setWindowTitle(f"Парсинг данных FIFA 21 с {host}")  # Устанавливаем заголовок окна
        central_widget = QWidget(self)  # Создаём центральный виджет
        self.setCentralWidget(central_widget)  # Устанавливаем центральный виджет

        self.grid_layout = QGridLayout()  # Создаём QGridLayout
        central_widget.setLayout(self.grid_layout)  # Устанавливаем данное размещение в центральный виджет

        self.table = QTableWidget(self)  # Создаём таблицу

        # Устанавливаем заголовки таблицы
        self.header_labels = []
        self.table.setHorizontalHeaderLabels(self.header_labels)

        # Устанавливаем всплывающие подсказки на заголовки
        # self.table.horizontalHeaderItem(0).setToolTip("Column 1 ")
        # self.table.horizontalHeaderItem(1).setToolTip("Column 2 ")
        # self.table.horizontalHeaderItem(2).setToolTip("Column 3 ")

        self.grid_layout.addWidget(self.table, 0, 0)  # Добавляем таблицу в сетку

        bottom_widget = QWidget(self)  # Создаём нижнюю панель
        self.control_grid_layout = QGridLayout()  # Создаём QGridLayout для панели управления
        bottom_widget.setLayout(self.control_grid_layout)  # Устанавливаем данное размещение в нижний виджет

        # информационная строка 1
        self.resultLabel = QtWidgets.QLabel(central_widget)
        self.resultLabel.setText('информация')
        self.resultLabel.setMaximumSize(220, 23)
        self.resultLabel.setAlignment(Qt.AlignLeft)
        self.control_grid_layout.addWidget(self.resultLabel, 0, 1)  # Добавляем лабел в сетку

        # информационная строка 2 (кол матчей)
        self.m_resultLabel = QtWidgets.QLabel(central_widget)
        self.m_resultLabel.setText('number off matches (0)')
        self.m_resultLabel.setMaximumSize(180, 23)
        self.m_resultLabel.setAlignment(Qt.AlignLeft)
        self.control_grid_layout.addWidget(self.m_resultLabel, 0, 2)  # Добавляем лабел в сетку

        # информационная строка 3 (time)
        self.timeLabel = QtWidgets.QLabel(central_widget)
        self.timeLabel.setText('общее время 00:00:00')
        self.timeLabel.setMaximumSize(180, 23)
        self.timeLabel.setAlignment(Qt.AlignLeft)
        self.control_grid_layout.addWidget(self.timeLabel, 0, 3)  # Добавляем лабел в сетку

        # информационная строка 4 (slider)
        self.slider_info = QtWidgets.QLabel(central_widget)
        self.slider_info.setText('interval: 1')
        self.slider_info.setMaximumSize(90, 23)
        self.slider_info.setAlignment(Qt.AlignLeft)
        self.slider_info.setStyleSheet('border: 4mm ridge rgba(170, 50, 220, .6);')
        self.control_grid_layout.addWidget(self.slider_info, 0, 4)  # Добавляем лабел в сетку

        # interval slider
        self.i_slider = QSlider(Qt.Horizontal, self)
        self.i_slider.setToolTip('pars interval value')
        self.i_slider.setMaximumSize(200, 23)
        self.i_slider.setMinimum(5)
        self.i_slider.setMaximum(100)
        self.i_slider.setValue(10)
        self.i_slider.valueChanged[int].connect(self.slider_val_change)
        self.control_grid_layout.addWidget(self.i_slider, 0, 5)  # Добавляем слайдер в сетку

        # кнопка авто скролл
        self.scrollButton = QtWidgets.QPushButton(central_widget)
        self.scrollButton.setFixedSize(140, 50)
        self.scrollButton.setObjectName("scrollButton")
        self.scrollButton.setText('Auto Scroll off')
        self.scrollButton.setStyleSheet('background: rgb(240, 62, 48); color: white;')
        self.scrollButton.clicked.connect(self.auto_scroll)
        self.control_grid_layout.addWidget(self.scrollButton, 0, 6)  # Добавляем кнопку в сетку

        # кнопка стоп
        self.stopButton = QtWidgets.QPushButton(central_widget)
        self.stopButton.setGeometry(QtCore.QRect(0, 0, 75, 23))
        self.stopButton.setFixedSize(280, 50)
        self.stopButton.setObjectName("stopButton")
        self.stopButton.setText('Stop')
        self.stopButton.setStyleSheet('background: rgb(240, 62, 48); color: white;')
        self.stopButton.clicked.connect(self.stop_pars)
        self.control_grid_layout.addWidget(self.stopButton, 0, 7)  # Добавляем кнопку в сетку

        # кнопка старт
        self.startButton = QtWidgets.QPushButton(central_widget)
        self.startButton.setGeometry(QtCore.QRect(0, 0, 75, 23))
        self.startButton.setFixedSize(280, 50)
        self.startButton.setObjectName("startButton")
        self.startButton.setText('Start')
        self.startButton.setStyleSheet('background: rgb(51, 48, 240); color: white;')
        self.startButton.clicked.connect(self.run_app)
        self.control_grid_layout.addWidget(self.startButton, 0, 8)  # Добавляем кнопку в сетку

        self.grid_layout.addWidget(bottom_widget, 1, 0)  # Добавляем control panel в сетку

        t = AsyncProcess('запуск драйвера', self.get_page_driver, 1, 'driver_is_ready')
        self.process = None

    def auto_scroll(self):
        self.auto = not self.auto
        if self.auto:
            self.scrollButton.setText('Auto Scroll on')
            self.scrollButton.setStyleSheet('background: rgb(51, 48, 240); color: white;')
        else:
            self.scrollButton.setText('Auto Scroll off')
            self.scrollButton.setStyleSheet('background: rgb(240, 62, 48); color: white;')

    def slider_val_change(self, value):
        self.interval = value / 10
        self.slider_info.setText(f'interval: {self.interval}')

    def timer(self):
        sec_ = 0
        min_ = 0
        hour_ = 0
        while True:
            self.timeLabel.setText('общее время - ' + time_str(hour_) + ':' + time_str(min_)
                                   + ':' + time_str(sec_))
            if self.end:
                break
            sleep(1)
            sec_ += 1
            if sec_ == 60:
                sec_ = 0
                min_ += 1
                if min_ == 60:
                    min_ = 0
                    hour_ += 1

    def run_app(self):
        if not self.run and self.driver:
            beep()
            self.start_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S").replace('/', '-').replace(' ', '_').replace(
                ':', '-')
            subprocess.call("powercfg -change -monitor-timeout-ac 0")
            subprocess.call("powercfg -change -disk-timeout-ac 0")
            subprocess.call("powercfg -change -standby-timeout-ac 0")
            self.end = False
            self.run = True
            self.total_list = {}  # последние значения total по матчам
            self.pars_rows = []  # массив данных спаршеных но не показаных на экран
            self.xlsx_data = []  # data to write xlsx
            self.end_match_check = {}  # number off last empty rows for matches
            self.row_color = 0  # 0 or 1 value permissible
            self.header_labels = []
            self.table.setColumnCount(0)
            self.table.setRowCount(0)
            self.resultLabel.setText('парсинг запущен')
            self.process = AsyncProcess('pars data', get_page_data, 1, 'end_pars')
            t = AsyncProcess('add rows async', self.add_row_async, 1, None)
            # start app timer async
            self.timeLabel.setText('общее время 00:00:00')
            t = AsyncProcess('timer', self.timer, 1, None)

    def stop_pars(self):
        if not self.run:
            return
        self.resultLabel.setText('парсинг остановлен')
        self.run = False

    def end_pars(self):
        self.run = False
        self.end = True
        self.resultLabel.setText('парсинг завершен')
        print('end')
        t = AsyncProcess('запись данных xlsx', write_xlsx_data, 1, 'end_after_save_data',
                         args=(self.start_time, self.xlsx_data,))

    def end_after_save_data(self):
        print('end pars after save data')
        beep()
        if self.close_app:
            subprocess.call("powercfg -SETACTIVE SCHEME_BALANCED")
            self.resultLabel.setText('app is closing')
            self.driver.quit()
            try:
                self.close()
            except:
                pass

    def closeEvent(self, event):
        print('close')
        self.resultLabel.setText('app is closing')
        self.close_app = True
        if not self.end:
            self.run = False
            event.ignore()
            return
        subprocess.call("powercfg -SETACTIVE SCHEME_BALANCED")
        self.driver.quit()
        event.accept()

    def get_page_driver(self, message=True):
        if message:
            self.resultLabel.setText('preparation')
        tables = None
        i = 0
        while not tables or len(tables) == 0:
            driver = parse_data_selenium(host)
            # link = driver.find_element_by_css_selector('menu.menu.js-header-menu li.menu__item:nth-child(2) a')
            # link.click()
            while not tables or len(tables) == 0:
                sleep(3)
                tables = driver.find_elements_by_css_selector('tbody.table__body')
                if host not in driver.current_url:
                    tables = None
                    driver.close()
                    break
                if tables and len(tables) > 0:
                    break
                if i >= 2:
                    driver.close()
                    break
                driver.execute_script("window.open('','_blank');")
                second_tab = driver.window_handles[1]
                driver.close()
                driver.switch_to.window(second_tab)
                driver.get(host)
                i += 1
        self.driver = driver

    def driver_is_ready(self):
        beep()
        self.resultLabel.setText('ready to start')

    def reload_driver(self, text):
        beep()
        if self.driver:
            self.driver.quit()
        self.resultLabel.setText(text)
        self.get_page_driver(False)
        self.resultLabel.setText('парсинг запущен')

    def set_num_matches(self, num):
        self.m_resultLabel.setText(f'number off matches ({num})')

    def set_color_to_row(self, row, col_count):
        if self.row_color == 0:
            self.row_color = 1
        else:
            self.row_color = 0
        for j in range(col_count - 1):
            if self.table.item(row, j).background() != self.grn_color:
                if self.row_color == 0:
                    self.table.item(row, j).setBackground(self.r_color)
                if j != 0 and (j + 1) % 4 == 0:
                    if '+' in self.table.item(row, j).text():
                        self.table.item(row, j).setBackground(self.grn_color)
                    else:
                        self.table.item(row, j).setBackground(self.red_color)
                if j != 0 and (j + 1) % 2 == 0 and (j + 1) % 4 != 0 and '!' in self.table.item(row, j).text():
                    self.table.item(row, j).setBackground(self.org_color)

    def get_interval(self, val_1, val_2):
        if not val_2 or val_2 == '' or val_1 == val_2.text():
            return '00:00'
        v_1 = self.time_to_sec(val_1)
        v_2 = self.time_to_sec(val_2.text())
        v = None
        if v_1 and v_2:
            v = v_1 - v_2
        val = self.sec_to_time(v)
        return val

    def sec_to_time(self, v):
        if not v:
            return '00:00'
        v_1 = str(math.trunc(v / 60))
        if len(v_1) == 1:
            v_1 = '0' + v_1
        v_2 = str(v % 60)
        if len(v_2) == 1:
            v_2 = '0' + v_2
        return v_1 + ':' + v_2

    def time_to_sec(self, str_):
        val = None
        if str_:
            v = str_.split(':')
            if len(v) >= 2:
                val = int(v[0]) * 60 + int(v[1])
        return val

    def add_row_data(self, data):
        self.pars_rows.append(data)

    def add_row_async(self):
        while self.run:
            sleep(self.interval / 2)
            if len(self.pars_rows) > 0:
                row = self.pars_rows[0]
                self.pars_rows.remove(self.pars_rows[0])
                t = threading.Thread(target=self.add_row, args=(row,))
                t.start()
                t.join()

    def check_table_size(self):
        cols_for_del = []
        i = 0
        for h in self.header_labels:
            if h and self.end_match_check.get(h) and self.end_match_check.get(h) > 10:
                for j in range(4):
                    cols_for_del.append(i + j)
            i += 1
        # for col in reversed(cols_for_del):
        #     self.table.removeColumn(col)
        #     self.header_labels.remove(self.header_labels[col])
        #     self.table.setHorizontalHeaderLabels(self.header_labels)

        # if self.table.rowCount() >= 2000:
        #     self.table.removeRow(0)

    def save_xlsx_row(self, row_items):
        xlsx_data = {}
        i = 0
        for h in self.header_labels:
            if h:
                x_data = []
                for j in range(4):
                    x_data.append(row_items[i + j])
                xlsx_data.update({h: x_data})
            i += 1
        self.xlsx_data.append(xlsx_data)
        if len(self.xlsx_data) >= 1000:
            t = AsyncProcess('запись данных xlsx', write_xlsx_data, 1, None, args=(self.start_time, self.xlsx_data,))
            self.start_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S").replace('/', '-').replace(' ', '_').replace(
                ':', '-')
            self.xlsx_data = []

    def append_match_data_to_row(self, row, row_items, data, j, header):
        # time interval
        val = '00:00'
        if row > 0:
            val = self.get_interval(data.get('time'), self.table.item(row - 1, j))
        # 1 column
        row_items.append(data.get('time'))
        # 2 column
        if data.get('total') != self.total_list.get(header):
            row_items.append(data.get('total') + ' !')
            self.total_list.update({header: data.get('total')})
        else:
            row_items.append(data.get('total'))
        # 3 column
        row_items.append(data.get('score'))
        # 4 column
        if val == '00:00':
            last = 2
            try:
                last = int(self.table.item(row - 1, j + 3).text()) + 1
            except:
                pass
            finally:
                row_items.append(str(last))
            if last > 2:
                self.table.setItem(row - 1, j + 3, QTableWidgetItem(''))
                self.table.item(row - 1, j + 3).setBackground(self.red_color)
        else:
            row_items.append(f'+{val} ')

    def add_row(self, row_data):
        self.check_table_size()
        m_num = len(row_data)
        row = self.table.rowCount()
        row_count = self.table.rowCount() + 1
        if len(self.header_labels) > 0:
            del self.header_labels[-1]
        j = 0
        row_items = []
        for h in self.header_labels:
            if h is not None:
                data = None
                try:
                    data = row_data.pop(h)
                except:
                    pass
                if data:
                    try:
                        self.end_match_check.pop(h)
                    except:
                        pass
                    self.append_match_data_to_row(row, row_items, data, j, h)
                    j += 4
                else:
                    n = self.end_match_check.get(h)
                    if n is None:
                        n = 0
                    self.end_match_check.update({h: n + 1})
                    for i in range(4):
                        row_items.append('')
                    j += 4
        for header in row_data.keys():
            self.header_labels.append(header)
            self.header_labels.append(None)
            self.header_labels.append(None)
            self.header_labels.append(None)
            # self.header_labels.append('T')
            # self.header_labels.append('S')
            # self.header_labels.append('I')
            data = row_data[header]
            self.append_match_data_to_row(row, row_items, data, j, header)
            j += 4
        # добавлем пустой столбец вконце таблицы
        self.header_labels.append(None)
        while len(row_items) < len(self.header_labels) - 1:
            row_items.append(None)
        col_count = len(self.header_labels)
        self.table.setColumnCount(col_count)
        self.table.setHorizontalHeaderLabels(self.header_labels)
        self.table.insertRow(row_count)
        self.table.setRowCount(row_count)
        self.set_num_matches(m_num)
        # записываем данные для сохранки в xlsx
        self.save_xlsx_row(row_items)
        # заполняем текущую строку
        i = 0
        for item in row_items:
            self.table.horizontalHeaderItem(i).setTextAlignment(Qt.AlignLeft)
            self.table.setItem(row, i, QTableWidgetItem(item))
            # if '+' in item:
            #     self.table.item(row, i).setBackground(self.grn_color)
            i += 1
        self.set_color_to_row(row, col_count)
        # делаем ресайз колонок по содержимому
        self.table.resizeColumnsToContents()
        # scroll to bottom row
        if self.auto and row_count > 17:
            self.table.scrollToBottom()
        # scroll to last column
        if self.auto and col_count > 20:
            self.table.scrollToItem(self.table.item(row, col_count - 2))


def main():
    global app_window
    app = QApplication(sys.argv)
    app_window = MainWindow()
    app_window.show()
    sys.exit(app.exec())


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    set_variables()
    main()
